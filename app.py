import streamlit as st
import requests
import base64
from PIL import Image
import pandas as pd
import io
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from tehlike_veritabani import (
    TEHLIKE_VERITABANI, VARSAYILAN_TEHLIKE,
    fine_kinney_skor, fine_kinney_seviye,
    l_matris_skor, l_matris_seviye
)

# Sayfa ayarları
st.set_page_config(
    page_title="İş Güvenliği Risk Değerlendirme Sistemi",
    page_icon="🦺",
    layout="wide"
)

# Session state
if "isyeri_bilgileri" not in st.session_state:
    st.session_state.isyeri_bilgileri = {}
if "analiz_sonucu" not in st.session_state:
    st.session_state.analiz_sonucu = None

# Başlık
st.title("🦺 İş Güvenliği Risk Değerlendirme Sistemi")
st.caption("Fine-Kinney ve L Matris (5x5) yöntemleri ile profesyonel risk analizi")

# Sidebar
with st.sidebar:
    st.header("⚙️ Ayarlar")
    api_key = st.text_input("Roboflow API Key", type="password")
    
    st.markdown("---")
    risk_method = st.radio(
        "📊 Risk Değerlendirme Yöntemi:",
        ["Fine-Kinney", "L Matris (5x5)", "Her İkisi"]
    )
    
    st.markdown("---")
    st.info("ℹ️ Roboflow AI ile iş güvenliği eksikliklerini tespit eder ve otomatik risk analizi yapar.")

# === SEKMELER ===
tab1, tab2, tab3 = st.tabs(["📋 1. İşyeri Bilgileri", "📸 2. Fotoğraf Analizi", "📑 3. Rapor"])

# === SEKME 1: İŞYERİ BİLGİLERİ ===
with tab1:
    st.subheader("İşyeri Bilgileri")
    st.caption("Risk değerlendirmesi raporunda kullanılacak işyeri bilgilerini doldurunuz.")
    
    col1, col2 = st.columns(2)
    
    with col1:
        isyeri_adi = st.text_input("İşyeri Ünvanı *", value=st.session_state.isyeri_bilgileri.get("isyeri_adi", ""))
        isyeri_adres = st.text_area("İşyeri Adresi *", value=st.session_state.isyeri_bilgileri.get("isyeri_adres", ""), height=80)
        sgk_no = st.text_input("SGK Sicil No", value=st.session_state.isyeri_bilgileri.get("sgk_no", ""))
        faaliyet_alani = st.text_input("Faaliyet Alanı", value=st.session_state.isyeri_bilgileri.get("faaliyet_alani", ""))
        calisan_sayisi = st.number_input("Çalışan Sayısı", min_value=1, value=st.session_state.isyeri_bilgileri.get("calisan_sayisi", 10))
    
    with col2:
        tehlike_sinifi = st.selectbox(
            "Tehlike Sınıfı",
            ["Az Tehlikeli", "Tehlikeli", "Çok Tehlikeli"],
            index=["Az Tehlikeli", "Tehlikeli", "Çok Tehlikeli"].index(
                st.session_state.isyeri_bilgileri.get("tehlike_sinifi", "Tehlikeli")
            )
        )
        analiz_tarihi = st.date_input("Analiz Tarihi", value=datetime.now().date())
        
        # Geçerlilik tarihi tehlike sınıfına göre
        gecerlilik_yil = {"Çok Tehlikeli": 2, "Tehlikeli": 4, "Az Tehlikeli": 6}
        gecerlilik_tarihi = analiz_tarihi.replace(year=analiz_tarihi.year + gecerlilik_yil[tehlike_sinifi])
        st.info(f"📅 Geçerlilik Tarihi: **{gecerlilik_tarihi.strftime('%d.%m.%Y')}** ({gecerlilik_yil[tehlike_sinifi]} yıl)")
        
        revizyon_no = st.text_input("Revizyon No", value=st.session_state.isyeri_bilgileri.get("revizyon_no", "0"))
    
    st.markdown("---")
    st.subheader("👥 Risk Değerlendirme Ekibi")
    
    col3, col4, col5 = st.columns(3)
    with col3:
        isveren = st.text_input("İşveren / İşveren Vekili", value=st.session_state.isyeri_bilgileri.get("isveren", ""))
        isg_uzmani = st.text_input("İş Güvenliği Uzmanı", value=st.session_state.isyeri_bilgileri.get("isg_uzmani", ""))
    with col4:
        isyeri_hekimi = st.text_input("İşyeri Hekimi", value=st.session_state.isyeri_bilgileri.get("isyeri_hekimi", ""))
        calisan_temsilcisi = st.text_input("Çalışan Temsilcisi", value=st.session_state.isyeri_bilgileri.get("calisan_temsilcisi", ""))
    with col5:
        destek_eleman = st.text_input("Destek Eleman", value=st.session_state.isyeri_bilgileri.get("destek_eleman", ""))
        bilgi_sahibi = st.text_input("Bilgi Sahibi Çalışan", value=st.session_state.isyeri_bilgileri.get("bilgi_sahibi", ""))
    
    if st.button("💾 Bilgileri Kaydet", type="primary", use_container_width=True):
        if not isyeri_adi or not isyeri_adres:
            st.error("⚠️ İşyeri Ünvanı ve Adresi zorunludur!")
        else:
            st.session_state.isyeri_bilgileri = {
                "isyeri_adi": isyeri_adi,
                "isyeri_adres": isyeri_adres,
                "sgk_no": sgk_no,
                "faaliyet_alani": faaliyet_alani,
                "calisan_sayisi": calisan_sayisi,
                "tehlike_sinifi": tehlike_sinifi,
                "analiz_tarihi": analiz_tarihi.strftime("%d.%m.%Y"),
                "gecerlilik_tarihi": gecerlilik_tarihi.strftime("%d.%m.%Y"),
                "revizyon_no": revizyon_no,
                "isveren": isveren,
                "isg_uzmani": isg_uzmani,
                "isyeri_hekimi": isyeri_hekimi,
                "calisan_temsilcisi": calisan_temsilcisi,
                "destek_eleman": destek_eleman,
                "bilgi_sahibi": bilgi_sahibi
            }
            st.success("✅ İşyeri bilgileri kaydedildi! Şimdi 'Fotoğraf Analizi' sekmesine geçebilirsiniz.")

# === SEKME 2: FOTOĞRAF ANALİZİ ===
with tab2:
    if not st.session_state.isyeri_bilgileri:
        st.warning("⚠️ Önce '1. İşyeri Bilgileri' sekmesinden bilgileri girip kaydetmeniz gerekiyor!")
    else:
        st.subheader("📸 Fotoğraf Analizi")
        st.caption(f"İşyeri: **{st.session_state.isyeri_bilgileri['isyeri_adi']}**")
        
        uploaded_file = st.file_uploader(
            "İş alanı fotoğrafı yükleyin",
            type=["jpg", "jpeg", "png"]
        )
        
        if uploaded_file is not None:
            col1, col2 = st.columns(2)
            with col1:
                st.markdown("**📷 Yüklenen Fotoğraf**")
                image = Image.open(uploaded_file)
                st.image(image, use_column_width=True)
            
            if not api_key:
                st.warning("⚠️ Sol menüden Roboflow API Key girin!")
            else:
                if st.button("🔍 Analiz Et", type="primary", use_container_width=True):
                    with st.spinner("🤖 Yapay zeka analiz yapıyor..."):
                        try:
                            image_rgb = image.convert("RGB")
                            buf = io.BytesIO()
                            image_rgb.save(buf, format="JPEG")
                            image_bytes = buf.getvalue()
                            image_base64 = base64.b64encode(image_bytes).decode("utf-8")
                            
                            url = "https://serverless.roboflow.com/infer/workflows/hseyins-workspace-vh1ss/detect-count-and-visualize"
                            payload = {
                                "api_key": api_key,
                                "inputs": {"image": {"type": "base64", "value": image_base64}}
                            }
                            
                            response = requests.post(url, json=payload, timeout=60)
                            response.raise_for_status()
                            result = response.json()
                            
                            outputs = result.get("outputs", [])
                            
                            # İşaretlenmiş görsel
                            with col2:
                                st.markdown("**🎯 Tespit Sonucu**")
                                if outputs:
                                    output = outputs[0]
                                    for key in ["annotated_image", "label_visualization", "detection_visualization"]:
                                        if key in output:
                                            img_obj = output[key]
                                            if isinstance(img_obj, dict) and "value" in img_obj:
                                                try:
                                                    img_data = base64.b64decode(img_obj["value"])
                                                    st.image(img_data, use_column_width=True)
                                                    break
                                                except:
                                                    pass
                            
                            # Tespit listesi
                            predictions = []
                            if outputs:
                                output = outputs[0]
                                for key in ["predictions", "model_predictions"]:
                                    if key in output:
                                        preds = output[key]
                                        if isinstance(preds, dict) and "predictions" in preds:
                                            predictions = preds["predictions"]
                                            break
                                        elif isinstance(preds, list):
                                            predictions = preds
                                            break
                            
                            st.session_state.analiz_sonucu = {
                                "predictions": predictions,
                                "raw_result": result
                            }
                            
                            if not predictions:
                                st.info("ℹ️ Fotoğrafta nesne tespit edilemedi.")
                            else:
                                st.success(f"✅ Toplam {len(predictions)} tespit yapıldı! 'Rapor' sekmesinden raporu görüntüleyebilirsiniz.")
                                
                                # Özet metrikler
                                col_a, col_b, col_c = st.columns(3)
                                col_a.metric("Toplam Tespit", len(predictions))
                                riskli = sum(1 for p in predictions if p.get("class", "").lower().startswith("no-"))
                                col_b.metric("⚠️ Riskli Durum", riskli)
                                col_c.metric("✅ Uyumlu Durum", len(predictions) - riskli)
                        
                        except requests.exceptions.HTTPError as e:
                            st.error(f"❌ API Hatası: {e.response.status_code}")
                        except Exception as e:
                            st.error(f"❌ Hata: {str(e)}")

# === SEKME 3: RAPOR ===
with tab3:
    if not st.session_state.isyeri_bilgileri:
        st.warning("⚠️ Önce işyeri bilgilerini girin!")
    elif not st.session_state.analiz_sonucu:
        st.warning("⚠️ Önce fotoğraf analizi yapın!")
    else:
        bilgiler = st.session_state.isyeri_bilgileri
        predictions = st.session_state.analiz_sonucu["predictions"]
        
        # Üst bilgi
        st.subheader("📑 Risk Değerlendirme Raporu")
        
        with st.container(border=True):
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**İşyeri:** {bilgiler['isyeri_adi']}")
                st.markdown(f"**Adres:** {bilgiler['isyeri_adres']}")
                st.markdown(f"**SGK Sicil No:** {bilgiler.get('sgk_no', '-')}")
                st.markdown(f"**Faaliyet Alanı:** {bilgiler.get('faaliyet_alani', '-')}")
            with col2:
                st.markdown(f"**Çalışan Sayısı:** {bilgiler['calisan_sayisi']}")
                st.markdown(f"**Tehlike Sınıfı:** {bilgiler['tehlike_sinifi']}")
                st.markdown(f"**Analiz Tarihi:** {bilgiler['analiz_tarihi']}")
                st.markdown(f"**Geçerlilik Tarihi:** {bilgiler['gecerlilik_tarihi']}")
        
        st.markdown("---")
        
        if not predictions:
            st.info("Tespit yok, rapor oluşturulamıyor.")
        else:
            # Risk listesi oluştur
            risk_listesi = []
            for idx, pred in enumerate(predictions, 1):
                sinif = pred.get("class", "bilinmeyen").lower()
                guven = pred.get("confidence", 0) * 100
                
                veri = TEHLIKE_VERITABANI.get(sinif, {**VARSAYILAN_TEHLIKE, "tehlike": f"Tespit: {sinif}"})
                
                # Mevcut skor
                fk_skor = fine_kinney_skor(veri["olasilik"], veri["siddet"], veri["frekans"])
                fk_seviye, fk_renk, fk_aciklama = fine_kinney_seviye(fk_skor)
                
                # DÖF sonrası skor
                fk_yeni = fine_kinney_skor(veri["yeni_olasilik"], veri["yeni_siddet"], veri["yeni_frekans"])
                fk_yeni_seviye, _, _ = fine_kinney_seviye(fk_yeni)
                
                # L Matris
                lm_skor = l_matris_skor(min(int(veri["olasilik"]), 5), min(int(veri["siddet"]/20)+1, 5))
                lm_seviye, lm_renk, _ = l_matris_seviye(lm_skor)
                
                satir = {
                    "No": idx,
                    "Bölüm/Faaliyet": veri["bolum"],
                    "Tehlike Kaynağı": veri["tehlike_kaynagi"],
                    "Tehlike": veri["tehlike"],
                    "Risk": veri["risk"],
                    "Mevcut Durum": veri["mevcut_durum"],
                    "Etkilenecek Kişiler": veri["etkilenecek"],
                    "O": veri["olasilik"],
                    "Ş": veri["siddet"],
                    "F": veri["frekans"],
                    "Risk Skoru": fk_skor,
                    "Risk Seviyesi": f"{fk_renk} {fk_seviye}",
                    "Alınacak Önlemler": veri["onlem"],
                    "Süreç Sorumlusu": bilgiler.get("isveren", ""),
                    "Termin": (datetime.now() + timedelta(days=90)).strftime("%d.%m.%Y"),
                    "Yeni O": veri["yeni_olasilik"],
                    "Yeni Ş": veri["yeni_siddet"],
                    "Yeni F": veri["yeni_frekans"],
                    "DÖF Sonrası Skor": fk_yeni,
                    "DÖF Sonrası Seviye": fk_yeni_seviye,
                    "Tespit Sınıfı": sinif,
                    "Güven (%)": f"{guven:.1f}"
                }
                
                if risk_method in ["L Matris (5x5)", "Her İkisi"]:
                    satir["LM Skor"] = lm_skor
                    satir["LM Seviye"] = f"{lm_renk} {lm_seviye}"
                
                risk_listesi.append(satir)
            
            df = pd.DataFrame(risk_listesi)
            
            # Tablo göster
            st.dataframe(df, use_container_width=True, hide_index=True, height=400)
            
            # İndirme butonları
            st.markdown("---")
            st.subheader("📥 Rapor İndirme")
            
            col_excel, col_pdf = st.columns(2)
            
            # ===== EXCEL =====
            with col_excel:
                excel_buffer = io.BytesIO()
                wb = Workbook()
                ws = wb.active
                ws.title = "Risk Analizi"
                
                # Stiller
                baslik_font = Font(name="Arial", size=14, bold=True)
                header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell_font = Font(name="Arial", size=9)
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                thin_border = Border(
                    left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin")
                )
                
                # Başlık
                ws.merge_cells("A1:L1")
                ws["A1"] = "RİSK DEĞERLENDİRME RAPORU - FİNE KİNNEY METODU"
                ws["A1"].font = baslik_font
                ws["A1"].alignment = center_align
                
                # İşyeri bilgileri
                ws["A3"] = "İşyeri Ünvanı:"
                ws["B3"] = bilgiler["isyeri_adi"]
                ws["A4"] = "Adres:"
                ws["B4"] = bilgiler["isyeri_adres"]
                ws["A5"] = "SGK Sicil No:"
                ws["B5"] = bilgiler.get("sgk_no", "")
                ws["A6"] = "Çalışan Sayısı:"
                ws["B6"] = bilgiler["calisan_sayisi"]
                ws["A7"] = "Tehlike Sınıfı:"
                ws["B7"] = bilgiler["tehlike_sinifi"]
                
                ws["E3"] = "Analiz Tarihi:"
                ws["F3"] = bilgiler["analiz_tarihi"]
                ws["E4"] = "Geçerlilik:"
                ws["F4"] = bilgiler["gecerlilik_tarihi"]
                ws["E5"] = "Revizyon No:"
                ws["F5"] = bilgiler["revizyon_no"]
                
                for row in range(3, 8):
                    ws[f"A{row}"].font = Font(bold=True)
                    ws[f"E{row}"].font = Font(bold=True)
                
                # Tablo başlıkları
                headers = ["No", "Bölüm/Faaliyet", "Tehlike Kaynağı", "Tehlike", "Risk",
                          "Etkilenecek", "O", "Ş", "F", "Risk Skoru", "Risk Seviyesi",
                          "Alınacak Önlemler", "Sorumlu", "Termin",
                          "Yeni O", "Yeni Ş", "Yeni F", "DÖF Sonrası Skor", "DÖF Sonrası Seviye"]
                
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=9, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                
                # Veri satırları
                for row_idx, satir in enumerate(risk_listesi, 10):
                    values = [
                        satir["No"], satir["Bölüm/Faaliyet"], satir["Tehlike Kaynağı"],
                        satir["Tehlike"], satir["Risk"], satir["Etkilenecek Kişiler"],
                        satir["O"], satir["Ş"], satir["F"], satir["Risk Skoru"],
                        satir["Risk Seviyesi"], satir["Alınacak Önlemler"],
                        satir["Süreç Sorumlusu"], satir["Termin"],
                        satir["Yeni O"], satir["Yeni Ş"], satir["Yeni F"],
                        satir["DÖF Sonrası Skor"], satir["DÖF Sonrası Seviye"]
                    ]
                    for col_idx, value in enumerate(values, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = cell_font
                        cell.border = thin_border
                        if col_idx in [1, 7, 8, 9, 10, 14, 15, 16, 17, 18]:
                            cell.alignment = center_align
                        else:
                            cell.alignment = left_align
                
                # Sütun genişlikleri
                col_widths = [5, 22, 25, 25, 30, 18, 5, 5, 5, 10, 18, 40, 18, 12, 6, 6, 6, 12, 18]
                for i, width in enumerate(col_widths, 1):
                    ws.column_dimensions[get_column_letter(i)].width = width
                
                # İmza alanı
                imza_row = len(risk_listesi) + 12
                ws.merge_cells(f"A{imza_row}:S{imza_row}")
                ws.cell(row=imza_row, column=1, value="ANALİZİ YAPANLAR").font = Font(bold=True, size=12)
                ws.cell(row=imza_row, column=1).alignment = center_align
                
                imza_baslik_row = imza_row + 2
                ws.cell(row=imza_baslik_row, column=1, value="İşveren / İşveren Vekili").font = Font(bold=True)
                ws.cell(row=imza_baslik_row, column=4, value="İş Güvenliği Uzmanı").font = Font(bold=True)
                ws.cell(row=imza_baslik_row, column=8, value="İşyeri Hekimi").font = Font(bold=True)
                ws.cell(row=imza_baslik_row, column=12, value="Çalışan Temsilcisi").font = Font(bold=True)
                
                ws.cell(row=imza_baslik_row + 1, column=1, value=bilgiler.get("isveren", ""))
                ws.cell(row=imza_baslik_row + 1, column=4, value=bilgiler.get("isg_uzmani", ""))
                ws.cell(row=imza_baslik_row + 1, column=8, value=bilgiler.get("isyeri_hekimi", ""))
                ws.cell(row=imza_baslik_row + 1, column=12, value=bilgiler.get("calisan_temsilcisi", ""))
                
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="📊 Excel Olarak İndir",
                    data=excel_buffer,
                    file_name=f"Risk_Analizi_{bilgiler['isyeri_adi'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            # ===== PDF =====
            with col_pdf:
                pdf_buffer = io.BytesIO()
                doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4),
                                       leftMargin=1*cm, rightMargin=1*cm,
                                       topMargin=1*cm, bottomMargin=1*cm)
                elements = []
                styles = getSampleStyleSheet()
                
                title_style = ParagraphStyle(
                    "Title", parent=styles["Heading1"],
                    fontSize=14, alignment=1, spaceAfter=12
                )
                elements.append(Paragraph("RİSK DEĞERLENDİRME RAPORU - FİNE KİNNEY METODU", title_style))
                
                # İşyeri bilgileri tablosu
                isyeri_data = [
                    ["İşyeri Ünvanı:", bilgiler["isyeri_adi"], "Analiz Tarihi:", bilgiler["analiz_tarihi"]],
                    ["Adres:", bilgiler["isyeri_adres"], "Geçerlilik:", bilgiler["gecerlilik_tarihi"]],
                    ["SGK No:", bilgiler.get("sgk_no", "-"), "Tehlike Sınıfı:", bilgiler["tehlike_sinifi"]],
                    ["Çalışan Sayısı:", str(bilgiler["calisan_sayisi"]), "Revizyon No:", bilgiler["revizyon_no"]]
                ]
                isyeri_table = Table(isyeri_data, colWidths=[3*cm, 8*cm, 3*cm, 5*cm])
                isyeri_table.setStyle(TableStyle([
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                    ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 4),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ]))
                elements.append(isyeri_table)
                elements.append(Spacer(1, 0.4*cm))
                
                # Risk tablosu
                pdf_headers = ["No", "Bölüm", "Tehlike", "Risk", "O", "Ş", "F", "Skor", "Seviye", "Önlem", "Yeni Skor"]
                pdf_data = [pdf_headers]
                
                for s in risk_listesi:
                    pdf_data.append([
                        str(s["No"]),
                        Paragraph(s["Bölüm/Faaliyet"][:40], styles["BodyText"]),
                        Paragraph(s["Tehlike"][:40], styles["BodyText"]),
                        Paragraph(s["Risk"][:50], styles["BodyText"]),
                        str(s["O"]), str(s["Ş"]), str(s["F"]),
                        str(s["Risk Skoru"]),
                        s["Risk Seviyesi"].replace("🟢", "").replace("🔵", "").replace("🟡", "").replace("🟠", "").replace("🔴", "").strip(),
                        Paragraph(s["Alınacak Önlemler"][:80], styles["BodyText"]),
                        f"{s['DÖF Sonrası Skor']}"
                    ])
                
                risk_table = Table(pdf_data, colWidths=[1*cm, 2.5*cm, 3*cm, 4*cm, 0.7*cm, 0.7*cm, 0.7*cm, 1*cm, 2*cm, 6*cm, 1.5*cm], repeatRows=1)
                risk_table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 8),
                    ("FONTSIZE", (0, 1), (-1, -1), 7),
                    ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                    ("ALIGN", (0, 1), (0, -1), "CENTER"),
                    ("ALIGN", (4, 1), (8, -1), "CENTER"),
                    ("ALIGN", (10, 1), (10, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
                ]))
                elements.append(risk_table)
                elements.append(Spacer(1, 0.5*cm))
                
                # İmza alanı
                imza_data = [
                    ["İşveren/İşveren Vekili", "İş Güvenliği Uzmanı", "İşyeri Hekimi", "Çalışan Temsilcisi"],
                    [bilgiler.get("isveren", ""), bilgiler.get("isg_uzmani", ""), bilgiler.get("isyeri_hekimi", ""), bilgiler.get("calisan_temsilcisi", "")],
                    ["İmza:", "İmza:", "İmza:", "İmza:"],
                    ["", "", "", ""]
                ]
                imza_table = Table(imza_data, colWidths=[6*cm, 6*cm, 6*cm, 6*cm], rowHeights=[0.7*cm, 0.7*cm, 0.5*cm, 1.5*cm])
                imza_table.setStyle(TableStyle([
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ]))
                elements.append(imza_table)
                
                doc.build(elements)
                pdf_buffer.seek(0)
                
                st.download_button(
                    label="📄 PDF Olarak İndir",
                    data=pdf_buffer,
                    file_name=f"Risk_Analizi_{bilgiler['isyeri_adi'].replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

# Footer
st.markdown("---")
st.caption("🦺 İş Güvenliği Risk Değerlendirme Sistemi v2.0 | Fine-Kinney & L Matris | Roboflow AI")
